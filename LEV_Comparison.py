import requests
import csv
import os
import datetime
import time
import logging
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from tenacity import retry, wait_exponential, stop_after_attempt, retry_if_exception_type

# --- Configuration and Constants ---
CISA_KEV_URL = "https://www.cisa.gov/sites/default/files/csv/known_exploited_vulnerabilities.csv"
NVD_API_BASE_URL = "https://services.nvd.nist.gov/rest/json/cves/2.0"
EPSS_API_BASE_URL = "https://api.first.org/data/v1/epss"
OUTPUT_EXCEL_FILE = "CVEs_KEV_EPSS_LEV.xlsx"
KEV_CSV_FILENAME_PREFIX = "known_exploited_vulnerabilities"
LOG_FILE = "vulnerability_analyzer.log"

# LEV Score Calculation Constants
EPSS_RELEVANCE_WINDOW_DAYS = 30
DATE_FORMAT = "%Y-%m-%d"

# --- Logging Setup ---
# Create a file handler
file_handler = logging.FileHandler(LOG_FILE)
file_handler.setLevel(logging.INFO) # Set level for file output
file_handler.setFormatter(logging.Formatter('%(asctime)s - %(levelname)s - %(message)s'))

# Create a stream handler (for console output)
stream_handler = logging.StreamHandler()
stream_handler.setLevel(logging.INFO) # Set level for console output
stream_handler.setFormatter(logging.Formatter('%(asctime)s - %(levelname)s - %(message)s'))

logging.basicConfig(
    level=logging.INFO, # Overall logging level
    handlers=[file_handler, stream_handler] # List of handlers
)

# --- Tenacity Retry Decorator ---
# Retry on common request exceptions and HTTP 429 (Too Many Requests)
# Wait exponentially between retries, up to 5 attempts.
@retry(
    wait=wait_exponential(multiplier=1, min=4, max=10),
    stop=stop_after_attempt(5),
    retry=retry_if_exception_type((
            requests.exceptions.RequestException,
            requests.exceptions.HTTPError
    ))
)
def make_throttled_request(url, params=None, headers=None):
    """
    Makes an HTTP GET request with throttling and retry logic.
    """
    logging.debug(f"Making request to: {url} with params: {params}")
    response = requests.get(url, params=params, headers=headers)
    response.raise_for_status() # Raise HTTPError for bad responses (4xx or 5xx)
    time.sleep(1) # Throttle: 1 request per second
    return response

# --- LEV Score Calculation Functions (Provided by User) ---
total_api_calls = 0 # This counter is for the LEV calculation's internal EPSS calls

def fetch_epss_score(cve_id, date_str):
    """
    Fetches EPSS score for a given CVE-ID and date.
    This function is called by calculate_lev and includes throttling and retries.
    """
    global total_api_calls
    total_api_calls += 1
    try:
        response = make_throttled_request(f"{EPSS_API_BASE_URL}?cve={cve_id}&date={date_str}")
        data = response.json().get("data") # EPSS API returns 'data' as a list [2]
        if not data: # Check if the list is empty
            logging.warning(f"EPSS API returned no data for {cve_id} on {date_str}. Response: {response.text}")
            return 0.0
        return float(data[0].get("epss", 0.0)) # Access the first item in the data list [2]
    except requests.exceptions.RequestException as e:
        logging.error(f"Failed to fetch EPSS for {cve_id} on {date_str} after retries: {e}")
        return 0.0
    except Exception as e:
        logging.error(f"Unexpected error fetching EPSS for {cve_id} on {date_str}: {e}")
        return 0.0

def weight_function(di, dn):
    delta = (dn - di).days
    return 1.0 if 0 <= delta < EPSS_RELEVANCE_WINDOW_DAYS else 0.0

def calculate_lev(cve_id, d0_str, dn_str, step_days=1):
    """
    Calculates the LEV score for a given CVE-ID.
    d0_str: CVE publication date (YYYY-MM-DD)
    dn_str: Current date (YYYY-MM-DD)
    """
    try:
        d0 = datetime.datetime.strptime(d0_str, DATE_FORMAT)
        dn = datetime.datetime.strptime(dn_str, DATE_FORMAT)
    except ValueError as e:
        logging.error(f"Date parsing error for {cve_id}: {e}. d0_str: {d0_str}, dn_str: {dn_str}")
        return None

    product = 1.0
    current_date = d0

    # Ensure we don't make excessive calls if d0 is very old
    # Limit the calculation window to EPSS_RELEVANCE_WINDOW_DAYS before dn
    # This ensures we only fetch EPSS scores relevant to the LEV window
    start_date_for_lev = max(d0, dn - datetime.timedelta(days=EPSS_RELEVANCE_WINDOW_DAYS))

    while current_date <= dn:
        # Only fetch EPSS if within the relevance window from the current_date to dn
        # This is a slight optimization based on the weight_function
        if (dn - current_date).days < EPSS_RELEVANCE_WINDOW_DAYS:
            date_str = current_date.strftime(DATE_FORMAT)
            epss_score = fetch_epss_score(cve_id, date_str)
            weight = weight_function(current_date, dn)
            term = 1.0 - (epss_score * weight)
            product *= term
        current_date += datetime.timedelta(days=step_days)

    lev_score = 1.0 - product
    return round(lev_score, 6)

# --- Data Fetching Functions ---

def fetch_nvd_data(cve_id):
    """
    Fetches CVSS v3.1 score, severity, and publication date from NVD API.
    """
    nvd_url = f"{NVD_API_BASE_URL}?cveId={cve_id}"
    try:
        response = make_throttled_request(nvd_url)
        data = response.json()

        cvss_score = "N/A"
        cvss_severity = "N/A"
        published_date = None

        vulnerabilities = data.get("vulnerabilities") # NVD API returns 'vulnerabilities' as a list [3]
        if not vulnerabilities: # Check if the list is empty
            logging.warning(f"NVD API returned no vulnerability data for {cve_id}. Response: {response.text}")
            return "N/A", "N/A", None

        # Access the first CVE object in the list, as cveId query should return one [1]
        cve_data = vulnerabilities[0].get("cve", {})

        # Extract published date
        published_date = cve_data.get("published")
        if published_date:
            # NVD date format is %Y-%m-%dT%H:%M:%S.sssZ, we need %Y-%m-%d [4, 5]
            published_date = published_date.split('T')[0]

        # Extract CVSS v3.1 data
        metrics = cve_data.get("metrics", {})
        cvss_v31_metrics = metrics.get("cvssMetricV31") # CVSS metrics are also a list [6, 7]

        # Prioritize 'Primary' type, otherwise take the first available [8]
        if cvss_v31_metrics:
            for metric in cvss_v31_metrics:
                if metric.get("type") == "Primary":
                    cvss_data = metric.get("cvssData", {})
                    cvss_score = cvss_data.get("baseScore", "N/A")
                    cvss_severity = cvss_data.get("baseSeverity", "N/A")
                    break
            else: # If no 'Primary' found, take the first one if available
                cvss_data = cvss_v31_metrics[0].get("cvssData", {}) # Access first element
                cvss_score = cvss_data.get("baseScore", "N/A")
                cvss_severity = cvss_data.get("baseSeverity", "N/A")

        if cvss_score == "N/A":
            logging.warning(f"CVSS v3.1 data not found for {cve_id} in NVD response.")
        if published_date is None:
            logging.warning(f"Published date not found for {cve_id} in NVD response.")

        return cvss_score, cvss_severity, published_date

    except requests.exceptions.RequestException as e:
        logging.error(f"Failed to fetch NVD data for {cve_id} after retries: {e}")
        return "Error", "Error", None
    except Exception as e:
        logging.error(f"Unexpected error fetching NVD data for {cve_id}: {e}")
        return "Error", "Error", None

def fetch_epss_current_data(cve_id):
    """
    Fetches the current EPSS score for a given CVE-ID.
    """
    epss_url = f"{EPSS_API_BASE_URL}?cve={cve_id}"
    try:
        response = make_throttled_request(epss_url)
        data = response.json().get("data") # EPSS API returns 'data' as a list [2]
        if not data: # Check if the list is empty
            logging.warning(f"Current EPSS data not found for {cve_id}.")
            return "N/A"
        return float(data[0].get("epss", 0.0)) # Access the first item in the data list [2]
    except requests.exceptions.RequestException as e:
        logging.error(f"Failed to fetch current EPSS data for {cve_id} after retries: {e}")
        return "Error"
    except Exception as e:
        logging.error(f"Unexpected error fetching current EPSS data for {cve_id}: {e}")
        return "Error"

# --- Main Script ---
def main():
    logging.info("Starting automated vulnerability analysis script.")

    # 1. Download KEV catalog
    timestamp = datetime.datetime.now().strftime("%Y%m%d-%H%M%S")
    timestamped_kev_filename = f"{KEV_CSV_FILENAME_PREFIX}_{timestamp}.csv"

    logging.info(f"Attempting to download KEV catalog from {CISA_KEV_URL}")
    try:
        response = make_throttled_request(CISA_KEV_URL)
        with open(timestamped_kev_filename, 'wb') as f:
            f.write(response.content)
        logging.info(f"KEV catalog downloaded and saved as {timestamped_kev_filename}")
    except Exception as e:
        logging.error(f"Failed to download KEV catalog: {e}")
        logging.error("Exiting script as KEV catalog is essential.")
        return

    # 2. Create Excel workbook and sheet
    wb = Workbook()
    ws = wb.active
    ws.title = "KEV_Analysis"
    logging.info(f"Created Excel workbook '{OUTPUT_EXCEL_FILE}' with sheet '{ws.title}'")

    # 3. Create columns and apply formatting
    headers = ["CVE ID", "CVSS Score", "CVSS Severity", "Current EPSS Score", "LEV Score"]
    ws.append(headers)
    for cell in ws[1]: # Apply bold font to headers
        cell.font = Font(bold=True)
    logging.info(f"Added headers to '{ws.title}' sheet.")

    # Prepare for auto-adjusting column widths
    column_widths = {i: len(header) for i, header in enumerate(headers, 1)}

    # 4. Process KEV entries
    cve_ids_to_process = []
    try:
        with open(timestamped_kev_filename, 'r', encoding='utf-8') as f:
            reader = csv.DictReader(f)
            for row in reader:
                # Assuming the CVE ID column in KEV CSV is named 'cveID'
                if 'cveID' in row:
                    cve_ids_to_process.append(row['cveID']) # Extract only the cveID string [User Instruction 7]
                else:
                    logging.warning(f"Row in KEV CSV missing 'cveID' column: {row}")
        logging.info(f"Found {len(cve_ids_to_process)} CVEs in the KEV catalog.")
    except Exception as e:
        logging.error(f"Failed to read CVE IDs from {timestamped_kev_filename}: {e}")
        logging.error("Exiting script as CVE data is essential.")
        return

    from tqdm import tqdm
    for cve_id_str in tqdm(cve_ids_to_process, desc="Processing CVEs"):
        logging.info(f"Processing CVE-ID: {cve_id_str}")

        cvss_score, cvss_severity, published_date = fetch_nvd_data(cve_id_str)
        epss_score_current = fetch_epss_current_data(cve_id_str)

        lev_score = "N/A"
        if published_date and published_date!= "N/A" and published_date!= "Error":
            dn_str = datetime.datetime.now().strftime(DATE_FORMAT)
            lev_score = calculate_lev(cve_id_str, published_date, dn_str)
            if lev_score is None:
                lev_score = "Error"
        else:
            logging.warning(f"Cannot calculate LEV for {cve_id_str}: Missing or invalid published date from NVD.")

        row_data = [cve_id_str, cvss_score, cvss_severity, epss_score_current, lev_score]
        ws.append(row_data)

        # Update column widths for auto-adjustment
        for i, cell_value in enumerate(row_data, 1):
            try:
                cell_length = len(str(cell_value))
                if column_widths[i] < cell_length:
                    column_widths[i] = cell_length
            except TypeError: # Handle non-string types gracefully
                pass

    # 5. Apply number formatting and auto-adjust column widths
    for col_idx, width in column_widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width + 2 # Add a little padding [7]

    # Apply number formatting to score columns
    for row_idx in range(2, ws.max_row + 1): # Start from second row (after headers)
        # CVSS Score (Column B)
        cvss_cell = ws.cell(row=row_idx, column=2)
        if isinstance(cvss_cell.value, (int, float)):
            cvss_cell.number_format = '0.0' # One decimal place

        # EPSS Score (Column D)
        epss_cell = ws.cell(row=row_idx, column=4)
        if isinstance(epss_cell.value, (int, float)):
            epss_cell.number_format = '0.0000' # Four decimal places

        # LEV Score (Column E)
        lev_cell = ws.cell(row=row_idx, column=5)
        if isinstance(lev_cell.value, (int, float)):
            lev_cell.number_format = '0.000000' # Six decimal places

    logging.info("Applied number formatting and auto-adjusted column widths.")

    # 6. Save the Excel file
    try:
        wb.save(OUTPUT_EXCEL_FILE)
        logging.info(f"Analysis complete. Results saved to '{OUTPUT_EXCEL_FILE}'")
    except Exception as e:
        logging.error(f"Failed to save Excel file '{OUTPUT_EXCEL_FILE}': {e}")

    logging.info("Script finished.")

if __name__ == "__main__":
    main()