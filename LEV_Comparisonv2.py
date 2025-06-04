# author Alexandre S. Cezar
# v1 - Data Retrieval and LEV calculation
# v2 - Paralell workers and API throttle handling

import requests
import csv
import os
import datetime
import time
import logging
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from tenacity import retry, wait_exponential, stop_after_attempt, retry_if_exception_type
from concurrent.futures import ThreadPoolExecutor, as_completed # Import for threading
from tqdm import tqdm # For progress bar

# --- Configuration ---
CISA_KEV_URL = "https://www.cisa.gov/sites/default/files/csv/known_exploited_vulnerabilities.csv"
NVD_API_BASE_URL = "https://services.nvd.nist.gov/rest/json/cves/2.0"
EPSS_API_BASE_URL = "https://api.first.org/data/v1/epss"
OUTPUT_EXCEL_FILE = "CVEs_KEV_EPSS_LEV.xlsx"
KEV_CSV_FILENAME_PREFIX = "known_exploited_vulnerabilities"
LOG_FILE = "vulnerability_analyzer.log"

# LEV Score Calculation Time Definition
EPSS_RELEVANCE_WINDOW_DAYS = 30
DATE_FORMAT = "%Y-%m-%d"

# --- Logging Setup ---
# Create a file handler
file_handler = logging.FileHandler(LOG_FILE)
file_handler.setLevel(logging.INFO) # Set level for file output
file_handler.setFormatter(logging.Formatter('%(asctime)s - %(levelname)s - %(message)s'))

# Create a stream handler (for console output)
stream_handler = logging.StreamHandler()
stream_handler.setLevel(logging.INFO) # Set level for console output
stream_handler.setFormatter(logging.Formatter('%(asctime)s - %(levelname)s - %(message)s'))

logging.basicConfig(
    level=logging.INFO, # Overall logging level
    handlers=[file_handler, stream_handler] # List of handlers
)

# --- Retry Definition ---
# Retry on common request exceptions and HTTP 429 (Too Many Requests)
# Wait exponentially between retries, up to 5 attempts.
@retry(
    wait=wait_exponential(multiplier=1, min=4, max=10),
    stop=stop_after_attempt(5),
    retry=retry_if_exception_type((
            requests.exceptions.RequestException,
            requests.exceptions.HTTPError
    ))
)
def make_throttled_request(url, params=None, headers=None):
    """
    Makes an HTTP GET request with throttling and retry logic.
    """
    logging.debug(f"Making request to: {url} with params: {params}")
    response = requests.get(url, params=params, headers=headers)
    response.raise_for_status() # Raise HTTPError for bad responses (4xx or 5xx)
    time.sleep(0.5) # Reduced throttling for concurrent requests, adjust as needed
    return response

# --- LEV Score Calculation Functions (That's the LEV Proposal Alghoritm) ---
total_api_calls = 0 # This counter is for the LEV calculation

def fetch_epss_score(cve_id, date_str):
    """
    Fetches EPSS score for a given CVE-ID and date.
    This function is called by calculate_lev and includes throttling and retries.
    """
    global total_api_calls
    total_api_calls += 1
    try:
        response = make_throttled_request(f"{EPSS_API_BASE_URL}?cve={cve_id}&date={date_str}")
        data = response.json().get("data")
        if not data:
            logging.warning(f"EPSS API returned no data for {cve_id} on {date_str}. Response: {response.text}")
            return 0.0
        # The data returned is a list of dictionaries, so we need to access the first element
        if isinstance(data, list) and data:
            return float(data[0].get("epss", 0.0))
        else:
            logging.warning(f"EPSS API data format unexpected for {cve_id} on {date_str}. Data: {data}")
            return 0.0
    except requests.exceptions.RequestException as e:
        logging.error(f"Failed to fetch EPSS for {cve_id} on {date_str} after retries: {e}")
        return 0.0
    except Exception as e:
        logging.error(f"Unexpected error fetching EPSS for {cve_id} on {date_str}: {e}")
        return 0.0

def weight_function(di, dn):
    delta = (dn - di).days
    return 1.0 if 0 <= delta < EPSS_RELEVANCE_WINDOW_DAYS else 0.0

def calculate_lev(cve_id, d0_str, dn_str, step_days=1):
    """
    Calculates the LEV score for a given CVE-ID.
    d0_str: CVE publication date (YYYY-MM-DD)
    dn_str: Current date (YYYY-MM-DD)
    """
    try:
        d0 = datetime.datetime.strptime(d0_str, DATE_FORMAT)
        dn = datetime.datetime.strptime(dn_str, DATE_FORMAT)
    except ValueError as e:
        logging.error(f"Date parsing error for {cve_id}: {e}. d0_str: {d0_str}, dn_str: {dn_str}")
        return None

    product = 1.0
    current_date = d0

    # Limit the calculation window to EPSS_RELEVANCE_WINDOW_DAYS before dn
    # This ensures we only fetch EPSS scores relevant to the LEV window
    start_date_for_lev = max(d0, dn - datetime.timedelta(days=EPSS_RELEVANCE_WINDOW_DAYS))

    # To parallelize EPSS score fetching within LEV calculation,
    # we can submit multiple fetch_epss_score tasks to a thread pool.
    # However, since `calculate_lev` is called *per CVE* in the main loop,
    # and EPSS scores are fetched sequentially for *each day* within a CVE's window,
    # the existing sequential call for `fetch_epss_score` within `calculate_lev`
    # is still a bottleneck and the code runs for a long time.
    # For simplicity and to avoid over-complicating the whole thing, I'll keep the daily EPSS fetches
    # within `calculate_lev` sequential for now, as the primary parallelization
    # will happen at the CVE level in `main`.
    # If this remains a bottleneck for *very long* CVE-ID histories, then
    # further parallelization *within* calculate_lev would be needed (more complex and I don't know how to do it).

    epss_dates_to_fetch = []
    temp_date = d0
    while temp_date <= dn:
        if (dn - temp_date).days < EPSS_RELEVANCE_WINDOW_DAYS:
            epss_dates_to_fetch.append(temp_date.strftime(DATE_FORMAT))
        temp_date += datetime.timedelta(days=step_days)

    # Fetch EPSS scores for all relevant dates for this CVE concurrently
    epss_scores_for_cve = {}
    with ThreadPoolExecutor(max_workers=5) as executor: # Use a smaller pool for internal LEV calculation
        future_to_date = {executor.submit(fetch_epss_score, cve_id, date_str): date_str for date_str in epss_dates_to_fetch}
        for future in as_completed(future_to_date):
            date_str = future_to_date[future]
            try:
                epss_scores_for_cve[date_str] = future.result()
            except Exception as exc:
                logging.error(f"EPSS score generation for {cve_id} on {date_str} generated an exception: {exc}")
                epss_scores_for_cve[date_str] = 0.0 # Default to 0 on error

    # Now calculate LEV using the fetched scores
    current_date_for_product = d0
    while current_date_for_product <= dn:
        date_str = current_date_for_product.strftime(DATE_FORMAT)
        epss_score = epss_scores_for_cve.get(date_str, 0.0) # Get pre-fetched score, default to 0
        weight = weight_function(current_date_for_product, dn)
        term = 1.0 - (epss_score * weight)
        product *= term
        current_date_for_product += datetime.timedelta(days=step_days)

    lev_score = 1.0 - product
    return round(lev_score, 6)

# --- Data Fetching Functions ---

def fetch_nvd_data(cve_id):
    """
    Fetches CVSS v3.1 score, severity, and publication date from NVD API.
    """
    nvd_url = f"{NVD_API_BASE_URL}?cveId={cve_id}"
    try:
        response = make_throttled_request(nvd_url)
        data = response.json()

        cvss_score = "N/A"
        cvss_severity = "N/A"
        published_date = None

        vulnerabilities = data.get("vulnerabilities")
        if not vulnerabilities:
            logging.warning(f"NVD API returned no vulnerability data for {cve_id}. Response: {response.text}")
            return "N/A", "N/A", None

        cve_data = vulnerabilities[0].get("cve", {})

        published_date = cve_data.get("published")
        if published_date:
            published_date = published_date.split('T')[0]

        metrics = cve_data.get("metrics", {})
        cvss_v31_metrics = metrics.get("cvssMetricV31")

        if cvss_v31_metrics:
            for metric in cvss_v31_metrics:
                if metric.get("type") == "Primary":
                    cvss_data = metric.get("cvssData", {})
                    cvss_score = cvss_data.get("baseScore", "N/A")
                    cvss_severity = cvss_data.get("baseSeverity", "N/A")
                    break
            else:
                if cvss_v31_metrics and len(cvss_v31_metrics) > 0: # Ensure there's at least one metric
                    cvss_data = cvss_v31_metrics[0].get("cvssData", {})
                    cvss_score = cvss_data.get("baseScore", "N/A")
                    cvss_severity = cvss_data.get("baseSeverity", "N/A")
        else: # Handle cases where cvss_v31_metrics is None or empty
            logging.warning(f"No CVSS v3.1 metrics found for {cve_id} in NVD response.")

        if cvss_score == "N/A":
            logging.warning(f"CVSS v3.1 data not found for {cve_id} in NVD response.")
        if published_date is None:
            logging.warning(f"Published date not found for {cve_id} in NVD response.")

        return cvss_score, cvss_severity, published_date

    except requests.exceptions.RequestException as e:
        logging.error(f"Failed to fetch NVD data for {cve_id} after retries: {e}")
        return "Error", "Error", None
    except Exception as e:
        logging.error(f"Unexpected error fetching NVD data for {cve_id}: {e}")
        return "Error", "Error", None

def fetch_epss_current_data(cve_id):
    """
    Fetches the current EPSS score for a given CVE-ID.
    """
    epss_url = f"{EPSS_API_BASE_URL}?cve={cve_id}"
    try:
        response = make_throttled_request(epss_url)
        data = response.json().get("data")
        if not data:
            logging.warning(f"Current EPSS data not found for {cve_id}.")
            return "N/A"
        # The data returned is a list of dictionaries, so we need to access the first element
        if isinstance(data, list) and data:
            return float(data[0].get("epss", 0.0))
        else:
            logging.warning(f"Current EPSS API data format unexpected for {cve_id}. Data: {data}")
            return "N/A"
    except requests.exceptions.RequestException as e:
        logging.error(f"Failed to fetch current EPSS data for {cve_id} after retries: {e}")
        return "Error"
    except Exception as e:
        logging.error(f"Unexpected error fetching current EPSS data for {cve_id}: {e}")
        return "Error"

# --- Main Script ---
def main():
    logging.info("Starting automated vulnerability analysis script.")

    # 1. Download KEV catalog
    timestamp = datetime.datetime.now().strftime("%Y%m%d-%H%M%S")
    timestamped_kev_filename = f"{KEV_CSV_FILENAME_PREFIX}_{timestamp}.csv"

    logging.info(f"Attempting to download KEV catalog from {CISA_KEV_URL}")
    try:
        response = make_throttled_request(CISA_KEV_URL)
        with open(timestamped_kev_filename, 'wb') as f:
            f.write(response.content)
        logging.info(f"KEV catalog downloaded and saved as {timestamped_kev_filename}")
    except Exception as e:
        logging.error(f"Failed to download KEV catalog: {e}")
        logging.error("Exiting script as KEV catalog is essential.")
        return

    # 2. Create Excel workbook and sheet
    wb = Workbook()
    ws = wb.active
    ws.title = "KEV_Analysis"
    logging.info(f"Created Excel workbook '{OUTPUT_EXCEL_FILE}' with sheet '{ws.title}'")

    # 3. Create columns and apply formatting
    headers = ["CVE ID", "CVSS Score", "CVSS Severity", "Current EPSS Score", "LEV Score"]
    ws.append(headers)
    for cell in ws[1]: # Apply bold font to headers
        cell.font = Font(bold=True)
    logging.info(f"Added headers to '{ws.title}' sheet.")

    # Prepare for auto-adjusting column widths
    column_widths = {i: len(header) for i, header in enumerate(headers, 1)}

    # 4. Process KEV entries
    cve_ids_to_process = []
    try:
        with open(timestamped_kev_filename, 'r', encoding='utf-8') as f:
            reader = csv.DictReader(f)
            # Skip the first row which is usually header in KEV CSV
            next(reader, None) # Skip header row if DictReader doesn't handle it
            for row in reader:
                if 'cveID' in row:
                    cve_ids_to_process.append(row['cveID'])
                else:
                    logging.warning(f"Row in KEV CSV missing 'cveID' column: {row}")
        logging.info(f"Found {len(cve_ids_to_process)} CVEs in the KEV catalog.")
    except Exception as e:
        logging.error(f"Failed to read CVE IDs from {timestamped_kev_filename}: {e}")
        logging.error("Exiting script as CVE data is essential.")
        return

    # Store results in a dictionary to maintain order and easily access data
    results = {}
    total_cves = len(cve_ids_to_process)
    logging.info(f"Starting parallel data fetching for {total_cves} CVEs...")

    # Define a helper function to encapsulate the per-CVE processing logic
    def process_single_cve(cve_id_str):
        # logging.info(f"Processing CVE-ID: {cve_id_str}") # Moved to debug to avoid overwhelming logs
        cvss_score, cvss_severity, published_date = fetch_nvd_data(cve_id_str)
        epss_score_current = fetch_epss_current_data(cve_id_str)

        lev_score = "N/A"
        if published_date and published_date != "N/A" and published_date != "Error":
            dn_str = datetime.datetime.now().strftime(DATE_FORMAT)
            lev_score = calculate_lev(cve_id_str, published_date, dn_str)
            if lev_score is None:
                lev_score = "Error"
        else:
            logging.warning(f"Cannot calculate LEV for {cve_id_str}: Missing or invalid published date from NVD.")
        return cve_id_str, cvss_score, cvss_severity, epss_score_current, lev_score, total_api_calls

    # Use ThreadPoolExecutor for concurrent API calls
    # Max workers can be adjusted based on network bandwidth and API rate limits.
    MAX_WORKERS = 20
    with ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
        # Submit all CVEs for processing
        future_to_cve = {executor.submit(process_single_cve, cve_id): cve_id for cve_id in cve_ids_to_process}

        # Use tqdm to show progress for completed futures
        for future in tqdm(as_completed(future_to_cve), total=total_cves, desc="Fetching CVE data"):
            cve_id_str = future_to_cve[future]
            try:
                # The result is the tuple returned by process_single_cve
                result_cve_id, cvss_score, cvss_severity, epss_score_current, lev_score, _ = future.result()
                results[result_cve_id] = [cvss_score, cvss_severity, epss_score_current, lev_score]
            except Exception as exc:
                logging.error(f"CVE {cve_id_str} generated an exception: {exc}")
                results[cve_id_str] = ["Error", "Error", "Error", "Error"] # Mark as error

    logging.info("Finished fetching all CVE data concurrently.")
    logging.info(f"Total internal EPSS API calls for LEV calculations: {total_api_calls}")

    # 5. Write results to Excel and apply formatting
    for cve_id_str in cve_ids_to_process: # Iterate in original order
        row_data = [cve_id_str] + results.get(cve_id_str, ["N/A", "N/A", "N/A", "N/A"]) # Use stored results
        ws.append(row_data)

        # Update column widths for auto-adjustment
        for i, cell_value in enumerate(row_data, 1):
            try:
                cell_length = len(str(cell_value))
                if column_widths[i] < cell_length:
                    column_widths[i] = cell_length
            except TypeError:
                pass

    # Apply column widths after all data is written
    for col_idx, width in column_widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width + 2

    # Apply number formatting to score columns
    for row_idx in range(2, ws.max_row + 1):
        cvss_cell = ws.cell(row=row_idx, column=2)
        if isinstance(cvss_cell.value, (int, float)):
            cvss_cell.number_format = '0.0'

        epss_cell = ws.cell(row=row_idx, column=4)
        if isinstance(epss_cell.value, (int, float)):
            epss_cell.number_format = '0.0000'

        lev_cell = ws.cell(row=row_idx, column=5)
        if isinstance(lev_cell.value, (int, float)):
            lev_cell.number_format = '0.000000'

    logging.info("Applied number formatting and auto-adjusted column widths.")

    # 6. Save the Excel file
    try:
        wb.save(OUTPUT_EXCEL_FILE)
        logging.info(f"Analysis complete. Results saved to '{OUTPUT_EXCEL_FILE}'")
    except Exception as e:
        logging.error(f"Failed to save Excel file '{OUTPUT_EXCEL_FILE}': {e}")

    logging.info("Script finished.")

if __name__ == "__main__":
    main()